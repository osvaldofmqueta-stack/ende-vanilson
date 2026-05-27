from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import RelatorioGerado
from django.utils import timezone
from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncMonth, TruncDay
from clientes.models import Cliente
from pagamentos.models import Pagamento, Fatura, Recarga
from equipamentos.models import Contador
from django.http import HttpResponse
import openpyxl
from io import BytesIO
from decimal import Decimal

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)


def is_admin_or_financeiro(user):
    return user.is_staff or (hasattr(user, 'perfil') and user.perfil.tipo_usuario in ['ADMIN', 'FINANCEIRO'])


# ─── Helpers ────────────────────────────────────────────────────────────────

AZUL_ESCURO  = colors.HexColor('#1a3a5c')
AZUL_CLARO   = colors.HexColor('#2d6a9f')
CINZA_HEADER = colors.HexColor('#e8f0f7')
CINZA_LINHA  = colors.HexColor('#f5f8fb')
VERMELHO     = colors.HexColor('#c0392b')
VERDE        = colors.HexColor('#1e8449')
LARANJA      = colors.HexColor('#d35400')

def _estilo_tabela_base(col_larguras, dados, zebra=True):
    estilo = [
        ('BACKGROUND',  (0, 0), (-1, 0),  AZUL_ESCURO),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  8),
        ('ALIGN',       (0, 0), (-1, 0),  'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING',  (0, 0), (-1, 0),  6),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('ALIGN',       (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING',  (0, 1), (-1, -1), 4),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#ccd9e8')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
            [CINZA_LINHA, colors.white] if zebra else [colors.white]),
    ]
    return TableStyle(estilo)


def _cabecalho_relatorio(styles, relatorio):
    elems = []
    titulo_style = ParagraphStyle(
        'TituloRel', parent=styles['Normal'],
        fontSize=18, textColor=AZUL_ESCURO,
        fontName='Helvetica-Bold', spaceAfter=2,
    )
    subtitulo_style = ParagraphStyle(
        'SubTituloRel', parent=styles['Normal'],
        fontSize=10, textColor=AZUL_CLARO,
        fontName='Helvetica', spaceAfter=2,
    )
    meta_style = ParagraphStyle(
        'MetaRel', parent=styles['Normal'],
        fontSize=8, textColor=colors.HexColor('#555555'),
        fontName='Helvetica',
    )

    elems.append(Paragraph("Sistema de Gestão de Energia", subtitulo_style))
    elems.append(Paragraph(relatorio.titulo, titulo_style))
    elems.append(HRFlowable(width="100%", thickness=2, color=AZUL_ESCURO, spaceAfter=4))

    periodo = f"{relatorio.periodo_inicio.strftime('%d/%m/%Y')} — {relatorio.periodo_fim.strftime('%d/%m/%Y')}"
    gerado  = relatorio.data_geracao.strftime('%d/%m/%Y %H:%M')
    elems.append(Paragraph(
        f"Período: <b>{periodo}</b>   |   Gerado por: <b>{relatorio.gerado_por}</b>   |   Em: <b>{gerado}</b>",
        meta_style
    ))
    elems.append(Spacer(1, 0.4 * cm))
    return elems


def _sem_dados(styles, msg="Nenhum dado encontrado para o período seleccionado."):
    p = ParagraphStyle('Aviso', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#888888'),
        fontName='Helvetica-Oblique', alignment=TA_CENTER)
    return [Spacer(1, 1*cm), Paragraph(msg, p)]


def _resumo_box(styles, itens):
    """Cria uma linha de resumo com KPIs."""
    data = [itens]
    col_w = [A4[0] / len(itens) - 1.2 * cm] * len(itens)
    t = Table([
        [Paragraph(f"<font size='16'><b>{v}</b></font><br/><font size='8' color='#555555'>{l}</font>", 
                   ParagraphStyle('kpi', parent=styles['Normal'], alignment=TA_CENTER))
         for l, v in itens]
    ], colWidths=col_w)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), CINZA_HEADER),
        ('BOX', (0,0), (-1,-1), 1, AZUL_CLARO),
        ('INNERGRID', (0,0), (-1,-1), 0.5, AZUL_CLARO),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    return [t, Spacer(1, 0.4*cm)]


# ─── Geradores por tipo ──────────────────────────────────────────────────────

def _pdf_clientes_ativos(elems, styles, relatorio):
    clientes = Cliente.objects.filter(
        data_cadastro__date__lte=relatorio.periodo_fim
    ).order_by('status', 'nome')

    ativos   = clientes.filter(status='ATIVO').count()
    inativos = clientes.filter(status='INATIVO').count()
    pre      = clientes.filter(tipo_cliente='PRE_PAGO').count()
    pos      = clientes.filter(tipo_cliente='POS_PAGO').count()

    elems += _resumo_box(styles, [
        ("Total de Clientes", clientes.count()),
        ("Ativos", ativos),
        ("Inativos", inativos),
        ("Pré-pago", pre),
        ("Pós-pago", pos),
    ])

    cabecalho = [['Nº Cliente', 'Nome', 'NIF', 'Tipo', 'Status', 'Data Cadastro']]
    dados = cabecalho + [
        [
            c.numero_cliente,
            Paragraph(c.nome, ParagraphStyle('n', fontSize=8, fontName='Helvetica')),
            c.nif,
            c.get_tipo_cliente_display(),
            c.get_status_display(),
            c.data_cadastro.strftime('%d/%m/%Y'),
        ]
        for c in clientes
    ]

    if len(dados) == 1:
        elems += _sem_dados(styles)
        return

    col_w = [2.5*cm, 5.5*cm, 2.5*cm, 2.2*cm, 2*cm, 2.5*cm]
    t = Table(dados, colWidths=col_w, repeatRows=1)
    t.setStyle(_estilo_tabela_base(col_w, dados))
    elems.append(t)


def _pdf_pagamentos(elems, styles, relatorio):
    faturas = Fatura.objects.filter(
        data_emissao__gte=relatorio.periodo_inicio,
        data_emissao__lte=relatorio.periodo_fim,
    ).select_related('cliente').order_by('-data_emissao')

    pagas     = faturas.filter(status='PAGO')
    pendentes = faturas.filter(status='PENDENTE')
    vencidas  = faturas.filter(status='VENCIDO')

    total_faturado = faturas.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')
    total_recebido = pagas.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')
    total_divida   = (pendentes | vencidas).aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')

    elems += _resumo_box(styles, [
        ("Total Faturado", f"{total_faturado:,.2f} Kz"),
        ("Recebido", f"{total_recebido:,.2f} Kz"),
        ("Em Dívida", f"{total_divida:,.2f} Kz"),
        ("Faturas Pagas", pagas.count()),
        ("Pendentes/Vencidas", pendentes.count() + vencidas.count()),
    ])

    cabecalho = [['Nº Fatura', 'Cliente', 'Emissão', 'Vencimento', 'Consumo (kWh)', 'Valor Total', 'Status']]
    status_cor = {'PAGO': VERDE, 'VENCIDO': VERMELHO, 'PENDENTE': LARANJA, 'CANCELADO': colors.grey}

    dados = cabecalho
    for f in faturas:
        dados.append([
            f.numero_fatura,
            Paragraph(f.cliente.nome, ParagraphStyle('n', fontSize=8, fontName='Helvetica')),
            f.data_emissao.strftime('%d/%m/%Y'),
            f.data_vencimento.strftime('%d/%m/%Y'),
            f"{f.consumo_kwh:.2f}",
            f"{f.valor_total:,.2f} Kz",
            f.get_status_display(),
        ])

    if len(dados) == 1:
        elems += _sem_dados(styles)
        return

    col_w = [2.8*cm, 4.5*cm, 2*cm, 2*cm, 2.5*cm, 2.7*cm, 2*cm]
    t = Table(dados, colWidths=col_w, repeatRows=1)
    estilo = _estilo_tabela_base(col_w, dados)

    for i, f in enumerate(faturas, start=1):
        cor = status_cor.get(f.status, colors.grey)
        estilo.add('TEXTCOLOR', (6, i), (6, i), cor)
        estilo.add('FONTNAME',  (6, i), (6, i), 'Helvetica-Bold')

    t.setStyle(estilo)
    elems.append(t)


def _pdf_financeiro(elems, styles, relatorio, agrupamento='mensal'):
    pagamentos = Pagamento.objects.filter(
        data_pagamento__date__gte=relatorio.periodo_inicio,
        data_pagamento__date__lte=relatorio.periodo_fim,
    )
    recargas = Recarga.objects.filter(
        data_recarga__date__gte=relatorio.periodo_inicio,
        data_recarga__date__lte=relatorio.periodo_fim,
        status='CONFIRMADO',
    )

    total_pag = pagamentos.aggregate(Sum('valor_pago'))['valor_pago__sum'] or Decimal('0')
    total_rec = recargas.aggregate(Sum('valor'))['valor__sum'] or Decimal('0')

    elems += _resumo_box(styles, [
        ("Pagamentos de Faturas", f"{total_pag:,.2f} Kz"),
        ("Recargas Confirmadas", f"{total_rec:,.2f} Kz"),
        ("Total Recebido", f"{total_pag + total_rec:,.2f} Kz"),
        ("Nº Transações", pagamentos.count() + recargas.count()),
    ])

    trunc = TruncMonth if agrupamento == 'mensal' else TruncDay
    label = 'Mês' if agrupamento == 'mensal' else 'Dia'
    date_fmt = '%m/%Y' if agrupamento == 'mensal' else '%d/%m/%Y'

    pag_grupo = (pagamentos
        .annotate(periodo=trunc('data_pagamento'))
        .values('periodo')
        .annotate(total=Sum('valor_pago'))
        .order_by('periodo'))

    rec_grupo = (recargas
        .annotate(periodo=trunc('data_recarga'))
        .values('periodo')
        .annotate(total=Sum('valor'))
        .order_by('periodo'))

    merged = {}
    for row in pag_grupo:
        k = row['periodo']
        merged.setdefault(k, {'pag': Decimal('0'), 'rec': Decimal('0')})
        merged[k]['pag'] += row['total'] or Decimal('0')
    for row in rec_grupo:
        k = row['periodo']
        merged.setdefault(k, {'pag': Decimal('0'), 'rec': Decimal('0')})
        merged[k]['rec'] += row['total'] or Decimal('0')

    cabecalho = [[label, 'Pagamentos Faturas', 'Recargas', 'Total']]
    dados = cabecalho + [
        [
            k.strftime(date_fmt),
            f"{v['pag']:,.2f} Kz",
            f"{v['rec']:,.2f} Kz",
            f"{v['pag']+v['rec']:,.2f} Kz",
        ]
        for k, v in sorted(merged.items())
    ]

    if len(dados) == 1:
        elems += _sem_dados(styles)
        return

    col_w = [3*cm, 5*cm, 5*cm, 5*cm]
    t = Table(dados, colWidths=col_w, repeatRows=1)
    estilo = _estilo_tabela_base(col_w, dados)
    for i in range(1, len(dados)):
        estilo.add('ALIGN', (1, i), (-1, i), 'RIGHT')
        estilo.add('FONTNAME', (3, i), (3, i), 'Helvetica-Bold')
    t.setStyle(estilo)
    elems.append(t)


def _pdf_consumo_mensal(elems, styles, relatorio):
    faturas = Fatura.objects.filter(
        data_emissao__gte=relatorio.periodo_inicio,
        data_emissao__lte=relatorio.periodo_fim,
    ).select_related('cliente', 'contador').order_by('cliente__nome', '-data_emissao')

    total_kwh = faturas.aggregate(Sum('consumo_kwh'))['consumo_kwh__sum'] or Decimal('0')
    total_val = faturas.aggregate(Sum('valor_total'))['valor_total__sum'] or Decimal('0')

    elems += _resumo_box(styles, [
        ("Total Consumo", f"{total_kwh:,.2f} kWh"),
        ("Total Faturado", f"{total_val:,.2f} Kz"),
        ("Nº Faturas", faturas.count()),
        ("Clientes Faturados", faturas.values('cliente').distinct().count()),
    ])

    cabecalho = [['Cliente', 'Contador', 'Período', 'Leit. Ant.', 'Leit. Act.', 'Consumo (kWh)', 'Valor (Kz)']]
    dados = cabecalho + [
        [
            Paragraph(f.cliente.nome, ParagraphStyle('n', fontSize=8, fontName='Helvetica')),
            f.contador.numero_serie if f.contador else '—',
            f.periodo_referencia,
            f"{f.leitura_anterior:.2f}",
            f"{f.leitura_atual:.2f}",
            f"{f.consumo_kwh:.2f}",
            f"{f.valor_total:,.2f}",
        ]
        for f in faturas
    ]

    if len(dados) == 1:
        elems += _sem_dados(styles)
        return

    col_w = [4*cm, 2.5*cm, 2.5*cm, 1.8*cm, 1.8*cm, 2.5*cm, 2.4*cm]
    t = Table(dados, colWidths=col_w, repeatRows=1)
    estilo = _estilo_tabela_base(col_w, dados)
    for i in range(1, len(dados)):
        estilo.add('ALIGN', (3, i), (-1, i), 'RIGHT')
    t.setStyle(estilo)
    elems.append(t)


def _pdf_consumo_area(elems, styles, relatorio):
    faturas = Fatura.objects.filter(
        data_emissao__gte=relatorio.periodo_inicio,
        data_emissao__lte=relatorio.periodo_fim,
    ).select_related('contador')

    area_map = {}
    for f in faturas:
        area = '—'
        if f.contador and f.contador.endereco_instalacao:
            partes = f.contador.endereco_instalacao.split(',')
            area = partes[-1].strip() if len(partes) > 1 else f.contador.endereco_instalacao[:30]
        area_map.setdefault(area, {'kwh': Decimal('0'), 'count': 0})
        area_map[area]['kwh']   += f.consumo_kwh or Decimal('0')
        area_map[area]['count'] += 1

    elems += _resumo_box(styles, [
        ("Áreas Identificadas", len(area_map)),
        ("Total Consumo", f"{sum(v['kwh'] for v in area_map.values()):,.2f} kWh"),
        ("Total Faturas", faturas.count()),
    ])

    cabecalho = [['Área / Zona', 'Nº Faturas', 'Consumo Total (kWh)', 'Consumo Médio (kWh)']]
    dados = cabecalho + sorted([
        [
            area,
            str(v['count']),
            f"{v['kwh']:,.2f}",
            f"{(v['kwh'] / v['count']):,.2f}" if v['count'] else '0.00',
        ]
        for area, v in area_map.items()
    ], key=lambda r: r[0])

    if len(dados) == 1:
        elems += _sem_dados(styles)
        return

    col_w = [7*cm, 3*cm, 5*cm, 5*cm]
    t = Table(dados, colWidths=col_w, repeatRows=1)
    estilo = _estilo_tabela_base(col_w, dados)
    for i in range(1, len(dados)):
        estilo.add('ALIGN', (1, i), (-1, i), 'RIGHT')
    t.setStyle(estilo)
    elems.append(t)


# ─── View principal ──────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin_or_financeiro)
def exportar_relatorio_pdf(request, pk):
    relatorio = get_object_or_404(RelatorioGerado, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{relatorio.id}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    elems  = _cabecalho_relatorio(styles, relatorio)

    tipo = relatorio.tipo_relatorio

    if tipo == 'CLIENTES_ATIVOS':
        _pdf_clientes_ativos(elems, styles, relatorio)
    elif tipo == 'PAGAMENTOS':
        _pdf_pagamentos(elems, styles, relatorio)
    elif tipo == 'FINANCEIRO_DIARIO':
        _pdf_financeiro(elems, styles, relatorio, agrupamento='diario')
    elif tipo == 'FINANCEIRO_MENSAL':
        _pdf_financeiro(elems, styles, relatorio, agrupamento='mensal')
    elif tipo == 'CONSUMO_MENSAL':
        _pdf_consumo_mensal(elems, styles, relatorio)
    elif tipo == 'CONSUMO_AREA':
        _pdf_consumo_area(elems, styles, relatorio)
    else:
        elems += _sem_dados(styles, f"Tipo de relatório '{relatorio.get_tipo_relatorio_display()}' sem dados específicos configurados.")

    # Rodapé com observações
    if relatorio.observacoes:
        obs_style = ParagraphStyle('Obs', parent=styles['Normal'],
            fontSize=8, textColor=colors.grey, fontName='Helvetica-Oblique')
        elems.append(Spacer(1, 0.5*cm))
        elems.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        elems.append(Spacer(1, 0.2*cm))
        elems.append(Paragraph(f"Observações: {relatorio.observacoes}", obs_style))

    doc.build(elems)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


# ─── Excel export (também melhorado) ─────────────────────────────────────────

@login_required
@user_passes_test(is_admin_or_financeiro)
def exportar_relatorio_excel(request, pk):
    relatorio = get_object_or_404(RelatorioGerado, pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    azul = PatternFill("solid", fgColor="1a3a5c")
    cinza = PatternFill("solid", fgColor="e8f0f7")
    bold_white = Font(bold=True, color="FFFFFF", size=11)
    bold_dark  = Font(bold=True, color="1a3a5c", size=10)
    normal     = Font(size=10)
    borda = Border(
        bottom=Side(style='thin', color='ccd9e8'),
        right=Side(style='thin', color='ccd9e8'),
    )

    def _header_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = bold_white
        c.fill = azul
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = borda
        return c

    def _data_cell(ws, row, col, value, is_even=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = normal
        if is_even:
            c.fill = cinza
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = borda
        return c

    # Meta
    ws['A1'] = "Sistema de Gestão de Energia"
    ws['A1'].font = Font(bold=True, size=14, color="1a3a5c")
    ws['A2'] = relatorio.titulo
    ws['A2'].font = Font(bold=True, size=12)
    ws['A3'] = f"Período: {relatorio.periodo_inicio.strftime('%d/%m/%Y')} a {relatorio.periodo_fim.strftime('%d/%m/%Y')}"
    ws['A4'] = f"Gerado por: {relatorio.gerado_por} em {relatorio.data_geracao.strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].font = ws['A4'].font = Font(size=10, color="555555")

    ws.append([])

    tipo = relatorio.tipo_relatorio

    if tipo == 'CLIENTES_ATIVOS':
        headers = ['Nº Cliente', 'Nome', 'NIF', 'BI', 'Tipo', 'Status', 'Telefone', 'Email', 'Data Cadastro']
        for col, h in enumerate(headers, 1):
            _header_cell(ws, ws.max_row + 1, col, h)
        clientes = Cliente.objects.filter(data_cadastro__date__lte=relatorio.periodo_fim).order_by('nome')
        for i, c in enumerate(clientes):
            r = ws.max_row + 1
            even = i % 2 == 1
            for col, val in enumerate([
                c.numero_cliente, c.nome, c.nif, c.bi,
                c.get_tipo_cliente_display(), c.get_status_display(),
                c.telefone, c.email or '', c.data_cadastro.strftime('%d/%m/%Y')
            ], 1):
                _data_cell(ws, r, col, val, even)

    elif tipo in ('PAGAMENTOS', 'CONSUMO_MENSAL', 'FINANCEIRO_DIARIO', 'FINANCEIRO_MENSAL'):
        faturas = Fatura.objects.filter(
            data_emissao__gte=relatorio.periodo_inicio,
            data_emissao__lte=relatorio.periodo_fim,
        ).select_related('cliente', 'contador')

        headers = ['Nº Fatura', 'Cliente', 'NIF', 'Período', 'Emissão', 'Vencimento',
                   'Consumo (kWh)', 'Valor Consumo (Kz)', 'Total (Kz)', 'Status']
        for col, h in enumerate(headers, 1):
            _header_cell(ws, ws.max_row + 1, col, h)
        for i, f in enumerate(faturas):
            r = ws.max_row + 1
            even = i % 2 == 1
            for col, val in enumerate([
                f.numero_fatura, f.cliente.nome, f.cliente.nif,
                f.periodo_referencia,
                f.data_emissao.strftime('%d/%m/%Y'),
                f.data_vencimento.strftime('%d/%m/%Y'),
                float(f.consumo_kwh), float(f.valor_consumo), float(f.valor_total),
                f.get_status_display()
            ], 1):
                _data_cell(ws, r, col, val, even)

    elif tipo == 'CONSUMO_AREA':
        headers = ['Área / Zona', 'Nº Faturas', 'Consumo Total (kWh)', 'Consumo Médio (kWh)']
        faturas = Fatura.objects.filter(
            data_emissao__gte=relatorio.periodo_inicio,
            data_emissao__lte=relatorio.periodo_fim,
        ).select_related('contador')
        area_map = {}
        for f in faturas:
            area = '—'
            if f.contador and f.contador.endereco_instalacao:
                partes = f.contador.endereco_instalacao.split(',')
                area = partes[-1].strip() if len(partes) > 1 else f.contador.endereco_instalacao[:30]
            area_map.setdefault(area, {'kwh': Decimal('0'), 'count': 0})
            area_map[area]['kwh'] += f.consumo_kwh or Decimal('0')
            area_map[area]['count'] += 1
        for col, h in enumerate(headers, 1):
            _header_cell(ws, ws.max_row + 1, col, h)
        for i, (area, v) in enumerate(sorted(area_map.items())):
            r = ws.max_row + 1
            even = i % 2 == 1
            media = float(v['kwh'] / v['count']) if v['count'] else 0
            for col, val in enumerate([area, v['count'], float(v['kwh']), media], 1):
                _data_cell(ws, r, col, val, even)

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 24

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{relatorio.id}.xlsx"'
    wb.save(response)
    return response


# ─── Outras views ─────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin_or_financeiro)
def relatorio_list(request):
    relatorios = RelatorioGerado.objects.all().order_by('-data_geracao')
    return render(request, 'relatorios/relatorio_list.html', {'relatorios': relatorios})


@login_required
@user_passes_test(is_admin_or_financeiro)
def gerar_relatorio_view(request):
    if request.method == 'POST':
        tipo   = request.POST.get('tipo_relatorio')
        inicio = request.POST.get('periodo_inicio')
        fim    = request.POST.get('periodo_fim')

        relatorio = RelatorioGerado.objects.create(
            titulo=f"Relatório — {dict(RelatorioGerado.TIPO_RELATORIO_CHOICES).get(tipo, tipo)}",
            tipo_relatorio=tipo,
            periodo_inicio=inicio,
            periodo_fim=fim,
            gerado_por=request.user.username,
            observacoes=f"Gerado automaticamente em {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        )
        return redirect('relatorio_list')

    return render(request, 'relatorios/relatorio_form.html', {
        'tipos': RelatorioGerado.TIPO_RELATORIO_CHOICES
    })


@login_required
@user_passes_test(is_admin_or_financeiro)
def estatisticas_gerais(request):
    total_clientes   = Cliente.objects.count()
    total_recebido   = Pagamento.objects.aggregate(Sum('valor_pago'))['valor_pago__sum'] or 0
    total_faturado   = Fatura.objects.aggregate(Sum('valor_total'))['valor_total__sum'] or 0
    faturas_vencidas = Fatura.objects.filter(status='VENCIDO').count()
    contadores_ativos = Contador.objects.filter(status='ATIVO').count()

    consumo_mensal = (Fatura.objects
        .annotate(mes=TruncMonth('data_emissao'))
        .values('mes')
        .annotate(total_kwh=Sum('consumo_kwh'), total_valor=Sum('valor_total'))
        .order_by('-mes')[:6])

    receita_mensal = (Pagamento.objects
        .annotate(mes=TruncMonth('data_pagamento'))
        .values('mes')
        .annotate(total_recebido=Sum('valor_pago'))
        .order_by('-mes')[:6])

    receita_diaria = (Pagamento.objects
        .annotate(dia=TruncDay('data_pagamento'))
        .values('dia')
        .annotate(total_recebido=Sum('valor_pago'))
        .order_by('-dia')[:15])

    data_inicio = request.GET.get('data_inicio')
    data_fim    = request.GET.get('data_fim')
    pagamentos_periodo = Pagamento.objects.all()

    if data_inicio and data_fim:
        pagamentos_periodo = pagamentos_periodo.filter(
            data_pagamento__date__range=[data_inicio, data_fim])

    total_periodo      = pagamentos_periodo.aggregate(Sum('valor_pago'))['valor_pago__sum'] or 0
    pagamentos_periodo = pagamentos_periodo.order_by('-data_pagamento')[:20]

    clientes_devedores = (Fatura.objects
        .filter(status__in=['PENDENTE', 'VENCIDO'])
        .values('cliente__nome', 'cliente__nif')
        .annotate(total_divida=Sum('valor_total'), faturas_count=Count('id'))
        .order_by('-total_divida')[:10])

    context = {
        'total_clientes': total_clientes,
        'total_recebido': total_recebido,
        'total_faturado': total_faturado,
        'faturas_vencidas': faturas_vencidas,
        'contadores_ativos': contadores_ativos,
        'consumo_mensal': consumo_mensal,
        'receita_mensal': receita_mensal,
        'receita_diaria': receita_diaria,
        'clientes_devedores': clientes_devedores,
        'pagamentos_periodo': pagamentos_periodo,
        'total_periodo': total_periodo,
        'filtros': {'inicio': data_inicio, 'fim': data_fim}
    }
    return render(request, 'relatorios/estatisticas.html', context)
